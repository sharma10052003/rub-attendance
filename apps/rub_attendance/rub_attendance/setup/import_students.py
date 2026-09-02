"""
Phase 1 student roster importer.

Targets the per-programme/section roster spreadsheets already in use at
Sherubtse College (see phase0/02-data-sources.md), e.g.:
    "1.BSc in DSDA Section A.xlsx"
    "3.BA in DCPM Year 1 Sem. I - Autumn 2026.xlsx"

Usage, from a real Frappe bench (this file cannot be executed on this
Windows machine — no bench/WSL/Docker here, see the chat for setup options):

    # Always dry-run first. Nothing is written to the database.
    bench --site <site> execute rub_attendance.setup.import_students.import_rosters \
        --kwargs "{'source_dir': '/path/to/Student Data 2026', 'dry_run': True}"

    # Once the dry-run report looks right, run for real. Safe to re-run —
    # upserts are keyed on student_id, so re-running never duplicates.
    bench --site <site> execute rub_attendance.setup.import_students.import_rosters \
        --kwargs "{'source_dir': '/path/to/Student Data 2026', 'dry_run': False}"

Before the first real run: create the Programme (and its Department) records
in Desk for every programme you're about to import. This importer will not
create them for you — which department owns which programme is a decision
for a person, not a guess baked into an import script. A roster whose
programme isn't found is reported as a skipped file, never invented.

Two more deliberate limits, both explained in phase0/02-data-sources.md:
- The "New admission with student numbers.xlsx" master file has a different
  column layout than the per-section rosters and is not handled here.
- Student names are not split into first/last. Bhutanese given names are
  frequently two or three words with no Western first/last structure, so
  the full name is stored as-is in Student.first_name and Student.last_name
  is left blank, rather than guessing a split that would corrupt the record.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

import frappe

STUDENT_ID_LENGTH = 8

# Known programme name fragments -> Programme.programme_code.
# Extend this as new programmes are piloted. A roster whose filename/title
# matches none of these is reported as an error, never guessed.
KNOWN_PROGRAMMES = {
	"DSDA": "DSDA",
	"DCPM": "DCPM",
	"EPS": "EPS",
	"MATHEMATICS": "MATH",
	"CHEMISTRY": "CHEM",
	"PHYSICS": "PHY",
	"LIFE SCIENCE": "LFSC",
}

HEADER_ALIASES = {
	"student_id": {"std. no.", "std no", "std no.", "student id", "student no", "student no."},
	"name": {"name", "student name"},
	"gender": {"gender", "sex"},
}

SECTION_PATTERN = re.compile(r"Section\s*([A-Za-z])", re.IGNORECASE)
YEAR_PATTERN = re.compile(r"(20\d{2})")
TERM_PATTERN = re.compile(r"(Autumn|Spring|Summer)", re.IGNORECASE)


@dataclass
class ImportReport:
	dry_run: bool = True
	files_processed: list = field(default_factory=list)
	files_skipped: list = field(default_factory=list)
	cohorts_created: list = field(default_factory=list)
	students_created: list = field(default_factory=list)
	students_updated: list = field(default_factory=list)
	row_errors: list = field(default_factory=list)
	warnings: list = field(default_factory=list)

	def as_dict(self):
		return {
			"dry_run": self.dry_run,
			"files_processed": self.files_processed,
			"files_skipped": self.files_skipped,
			"cohorts_created": self.cohorts_created,
			"students_created_count": len(self.students_created),
			"students_updated_count": len(self.students_updated),
			"students_created": self.students_created,
			"students_updated": self.students_updated,
			"row_errors": self.row_errors,
			"warnings": self.warnings,
		}


def identify_programme(text: str):
	text_upper = text.upper()
	for fragment, code in KNOWN_PROGRAMMES.items():
		if fragment in text_upper:
			return code
	return None


def parse_roster_identity(path: Path, title_text: str = ""):
	"""Return (identity_dict, warnings) or (None, errors) for a roster file.

	Matches against the filename AND the sheet's own title text, because real
	Sherubtse roster files are inconsistent about which one carries the year —
	e.g. "1.BSc in DSDA Section A.xlsx" has no year in its filename at all,
	but its sheet contains the string "Bachelor of Data Science and Data
	Analytics - 2026"."""
	text = f"{path.stem} {title_text}"

	programme_code = identify_programme(text)
	year_match = YEAR_PATTERN.search(text)

	errors = []
	if not programme_code:
		errors.append(
			f"Could not identify a known programme in filename '{path.name}'. "
			f"Add it to KNOWN_PROGRAMMES in import_students.py if this is a new programme."
		)
	if not year_match:
		errors.append(f"Could not find a 4-digit intake year (20xx) in filename '{path.name}'.")
	if errors:
		return None, errors

	section_match = SECTION_PATTERN.search(text)
	term_match = TERM_PATTERN.search(text)

	warnings = []
	if section_match:
		section = section_match.group(1).upper()
	else:
		section = "A"
		warnings.append(
			f"No 'Section X' found in '{path.name}' — assumed Section A "
			f"(treated as a single-section programme). Verify this is correct."
		)

	return {
		"programme_code": programme_code,
		"intake_year": int(year_match.group(1)),
		"section": section,
		"term": term_match.group(1).title() if term_match else None,
	}, warnings


def normalize_student_id(value):
	"""RUB student IDs are 8 digits (e.g. 07260124). Excel frequently reads a
	numeric-typed cell as an int and drops the leading zero — zero-pad it back."""
	if value is None:
		return None
	if isinstance(value, float):
		value = int(value)
	text = str(value).strip().replace(" ", "")
	if not text.isdigit():
		return None
	return text.zfill(STUDENT_ID_LENGTH)


def normalize_gender(value):
	if not value:
		return None
	text = str(value).strip().lower()
	if text in ("f", "female"):
		return "Female"
	if text in ("m", "male"):
		return "Male"
	return None


def read_roster_rows(path: Path):
	"""Returns (col_index, data_rows, title_text). title_text is every string
	cell found in rows before the header row — the sheet's title/caption, if
	any — used as a fallback source for programme/year/section identity."""
	from openpyxl import load_workbook

	workbook = load_workbook(path, read_only=True, data_only=True)
	try:
		sheet = workbook.active
		row_iter = sheet.iter_rows(values_only=True)

		col_index = {}
		title_parts = []
		for row in row_iter:
			normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
			matches = {}
			for key, aliases in HEADER_ALIASES.items():
				for idx, cell in enumerate(normalized):
					if cell in aliases:
						matches[key] = idx
						break
			if "name" in matches and "student_id" in matches:
				col_index = matches
				break

			for cell in row:
				if isinstance(cell, str) and len(cell.strip()) > 10:
					title_parts.append(cell.strip())

		if not col_index:
			raise ValueError(
				f"Could not find a header row with recognizable columns "
				f"(Std. No. / Name / Gender) in {path.name}"
			)

		data_rows = [row for row in row_iter if any(cell is not None for cell in row)]
		return col_index, data_rows, " ".join(title_parts)
	finally:
		workbook.close()


def get_or_create_cohort(programme: str, intake_year: int, section: str, dry_run: bool, report: ImportReport):
	existing = frappe.db.get_value(
		"Cohort",
		{"programme": programme, "intake_year": intake_year, "section": section},
		"name",
	)
	if existing:
		return existing

	if dry_run:
		report.cohorts_created.append(f"{programme} / {intake_year} / Section {section} (would be created)")
		return f"<pending:{programme}:{intake_year}:{section}>"

	doc = frappe.get_doc(
		{
			"doctype": "Cohort",
			"programme": programme,
			"intake_year": intake_year,
			"section": section,
		}
	).insert(ignore_permissions=True)
	frappe.db.commit()
	report.cohorts_created.append(doc.name)
	return doc.name


def upsert_student(student_id: str, name: str, gender, cohort_name: str, dry_run: bool, report: ImportReport):
	existing = frappe.db.exists("Student", student_id)

	if dry_run:
		if existing:
			report.students_updated.append(student_id)
		else:
			report.students_created.append(student_id)
		return

	if cohort_name.startswith("<pending:"):
		frappe.throw(
			"Internal error: attempted a real write against a not-yet-created cohort. "
			"This should never happen outside dry_run=True — please report it."
		)

	try:
		if existing:
			doc = frappe.get_doc("Student", student_id)
			doc.first_name = name
			if gender:
				doc.gender = gender
			doc.cohort = cohort_name
			doc.save(ignore_permissions=True)
			report.students_updated.append(student_id)
		else:
			doc = frappe.get_doc(
				{
					"doctype": "Student",
					"student_id": student_id,
					"first_name": name,
					"gender": gender,
					"cohort": cohort_name,
					"status": "Active",
				}
			).insert(ignore_permissions=True)
			report.students_created.append(student_id)
		frappe.db.commit()
	except Exception as e:
		frappe.db.rollback()
		report.row_errors.append(
			{"file": None, "row": None, "error": f"student {student_id}: {e}"}
		)


def import_rosters(source_dir: str, dry_run: bool = True) -> dict:
	report = ImportReport(dry_run=dry_run)
	source = Path(source_dir)
	if not source.exists():
		frappe.throw(f"Source directory not found: {source_dir}")

	for path in sorted(source.glob("*.xlsx")):
		if path.name.startswith("~$"):
			continue  # Excel lock file, not real data

		if "admission" in path.stem.lower():
			report.files_skipped.append(
				{
					"file": path.name,
					"reason": (
						"Admissions master file has a different column layout than the "
						"per-section rosters — not handled by this importer. See "
						"phase0/02-data-sources.md."
					),
				}
			)
			continue

		try:
			col_index, data_rows, title_text = read_roster_rows(path)
		except ValueError as e:
			report.files_skipped.append({"file": path.name, "reason": str(e)})
			continue

		identity, notes = parse_roster_identity(path, title_text)
		if identity is None:
			report.files_skipped.append({"file": path.name, "reason": "; ".join(notes)})
			continue
		for note in notes:
			report.warnings.append({"file": path.name, "warning": note})

		programme_name = frappe.db.get_value(
			"Programme", {"programme_code": identity["programme_code"]}, "name"
		)
		if not programme_name:
			report.files_skipped.append(
				{
					"file": path.name,
					"reason": (
						f"No Programme found with programme_code="
						f"{identity['programme_code']!r}. Create it (and its Department) "
						f"in Desk before importing this file."
					),
				}
			)
			continue

		cohort_name = get_or_create_cohort(
			programme_name, identity["intake_year"], identity["section"], dry_run, report
		)

		for row_number, row in enumerate(data_rows, start=1):
			student_id = normalize_student_id(row[col_index["student_id"]])
			name_raw = row[col_index["name"]]
			name = str(name_raw).strip() if name_raw else ""
			gender = normalize_gender(
				row[col_index["gender"]] if "gender" in col_index else None
			)

			if not student_id:
				report.row_errors.append(
					{
						"file": path.name,
						"row": row_number,
						"error": f"missing/invalid student ID: {row[col_index['student_id']]!r}",
					}
				)
				continue
			if not name:
				report.row_errors.append(
					{
						"file": path.name,
						"row": row_number,
						"error": f"missing name for student ID {student_id}",
					}
				)
				continue

			upsert_student(student_id, name, gender, cohort_name, dry_run, report)

		report.files_processed.append(path.name)

	return report.as_dict()


@frappe.whitelist()
def import_rosters_api(source_dir: str, dry_run: bool = True):
	"""Whitelisted wrapper so a future Desk button can trigger this without
	console access. Restricted to Registry/System Manager — the same people
	who can write Student records per phase0/05-permission-matrix.md."""
	if not (
		"System Manager" in frappe.get_roles()
		or "Registry" in frappe.get_roles()
	):
		frappe.throw("Not permitted", frappe.PermissionError)
	if isinstance(dry_run, str):
		dry_run = dry_run.lower() not in ("false", "0", "")
	return import_rosters(source_dir, dry_run=dry_run)
